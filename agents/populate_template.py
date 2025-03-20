# import openpyxl
# from io import BytesIO
# import streamlit as st

# def populate_template(master_df, mapping, template_file):
#     """Creates a new Excel file with mapped data from master_df while retaining template columns."""

#     if not isinstance(mapping, dict):
#         raise TypeError(f"Expected dictionary for mapping, but got {type(mapping)}")

#     try:
#         # Load the original template (preserve the columns only)
#         template_wb = openpyxl.load_workbook(template_file)
#         template_sheet = template_wb.active  # Assuming single-sheet templates

#         # Extract headers from the template (only column names, not data)
#         template_headers = [cell.value for cell in template_sheet[1] if cell.value]  # First row is header

#         # Create a new workbook
#         new_wb = openpyxl.Workbook()
#         new_sheet = new_wb.active
#         new_sheet.title = "Sheet1"  # Name of the sheet (optional)

#         # Write headers to the first row in the new sheet (using template headers)
#         new_sheet.append(template_headers)  # Write headers to the first row

#         # Populate the new sheet with data from master_df based on the mapping
#         for _, row in master_df.iterrows():
#             new_row = []
#             for template_col in template_headers:
#                 # Find the corresponding column in master_df using mapping
#                 csv_col = next((key for key, value in mapping.items() if value == template_col), None)
#                 if csv_col and csv_col in master_df.columns:
#                     new_row.append(row[csv_col])  # Add the value from master_df
#                 else:
#                     new_row.append("")  # If no match, add empty string
#             new_sheet.append(new_row)  # Write the row to the sheet

#         # Save the new workbook to a BytesIO object for streaming or saving
#         output = BytesIO()
#         new_wb.save(output)
#         output.seek(0)

#         # Optional: Display the first few rows of the new data in Streamlit
#         st.write(master_df.head(10))
#         st.write("✅ New Excel File Created Successfully!")

#         return output

#     except Exception as e:
#         st.error(f"⚠️ Error creating new Excel file: {e}")
#         return None
############################################################################################################
# import openpyxl
# from io import BytesIO
# import streamlit as st
# import openai
# import os
# from dotenv import load_dotenv
# import json

# load_dotenv()
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# # OpenAI API Key (Use Environment Variable Instead of Hardcoding)
# # OPENAI_API_KEY = "your-api-key"

# def get_ai_column_mapping(master_df, template_headers):
#     """Uses OpenAI to infer the best mapping between master_df columns and template headers."""
    
#     prompt =(
#         f"Given these CSV columns: {list(master_df.columns)}, map them to these template columns: {template_headers}. "
#         "Return the mapping as a **valid JSON object** with the format: "
#         '{ "mappings": { "CSV_column_name": "Template_header_name", ... } }. '
#         "Keep mappings as key in the response JSON object."
#         "- Match the best column names based on similarity and meaning."
#         "- If no exact match is found, provide the closest reasonable alternative."
#         "- Return a dictionary where the key is the column from master_df, and the value is the corresponding template header."
#     )
#     st.write("✅ Trying AI Column Mapping:")
#     response = openai.ChatCompletion.create(
#         model="gpt-4-turbo",
#         messages=[{"role": "system", "content": "You are an AI assistant that maps map CSV data to the Excel template in JSON format"},
#                   {"role": "user", "content": prompt}],
#         api_key=OPENAI_API_KEY
#     )
#     st.write("Response:",response)

#     try:
#         content = response["choices"][0]["message"]["content"]

#         # Extract JSON from response using regex or splitting
#         json_start = content.find("{")
#         json_end = content.rfind("}") + 1
#         json_str = content[json_start:json_end]  # Extract JSON part

#         # Parse JSON to dictionary
#         ai_mapping = json.loads(json_str)["mappings"]
#         st.write("✅ AI Column Mapping:", ai_mapping)
#         return ai_mapping
    
#     except Exception as e:
#         st.error(f"⚠️ Error parsing AI response: {e}")
#         return {}
        
#     #     st.write("✅ Trying AI Column Mapping:", )
#     #     ai_mapping = eval(response["choices"][0]["message"]["content"])["mappings"]  # Convert string response to dict
#     #     st.write("✅ AI Column Mapping:", ai_mapping)
#     #     return ai_mapping
#     # except Exception as e:
#     #     st.error(f"⚠️ Error in AI column mapping: {e}")
#     #     return {}

# def populate_template(master_df, template_file):
#     """Creates a new Excel file with AI-mapped data while retaining template columns."""

#     try:
#         # Load the original template
#         template_wb = openpyxl.load_workbook(template_file)
#         template_sheet = template_wb.active  # Assuming single-sheet templates
#         # Extract headers from the template
#         template_headers = [cell.value for cell in template_sheet[1] if cell.value]  # First row is header
        
#         print("master_df:",master_df.columns)
#         print("Template headers:",template_headers)
#         st.write("✅ Loading Excel Template:", template_file)
#         # Use OpenAI API to infer column mappings
#         mapping = get_ai_column_mapping(master_df, template_headers)
#         if not mapping:
#             st.error("⚠️ AI failed to generate a valid column mapping.")
#             return None

#         # Create a new workbook
#         new_wb = openpyxl.Workbook()
#         new_sheet = new_wb.active
#         new_sheet.title = "Sheet1"

#         # Write headers to the first row in the new sheet
#         new_sheet.append(template_headers)

#         # Populate the new sheet with mapped data
#         for _, row in master_df.iterrows():
#             new_row = []
#             for template_col in template_headers:
#                 # Find the corresponding column in master_df using AI-inferred mapping
#                 csv_col = next((key for key, value in mapping.items() if value == template_col), None)
#                 new_row.append(row[csv_col] if csv_col and csv_col in master_df.columns else "")  # Add value or empty
#             new_sheet.append(new_row)  # Write row to sheet

#         # Save the new workbook to a BytesIO object
#         output = BytesIO()
#         new_wb.save(output)
#         output.seek(0)

#         # Display output preview in Streamlit
#         st.write(master_df.head(10))
#         st.write("✅ New Excel File Created Successfully with AI-generated column mapping!")

#         return output

#     except Exception as e:
#         st.error(f"⚠️ Error creating new Excel file: {e}")
#         return None

############################################################################################################
import openpyxl
from io import BytesIO
import streamlit as st
import openai
import os
from dotenv import load_dotenv
import json
import pandas as pd

load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

def get_ai_transformed_data(master_df, template_headers):
    """Uses OpenAI to generate a transformed dataset matching the template."""
    
    # Convert DataFrame to JSON format to send structured data to AI
    sample_data = master_df.to_dict(orient="records")  # Send first 10 rows
    
    prompt = (
        f"Given this dataset (JSON format):\n{json.dumps(sample_data, indent=2)}\n\n"
        f"Transform it to match the following template columns: {template_headers}.\n"
        "- Perform necessary calculations and transformations.\n"
        "- Return only a JSON array where each object represents a row in the final dataset."
        "- Ensure correct mappings, format conversions, and required calculations."
        "- Do NOT return any extra text, only a valid JSON list of dictionaries."
    )

    st.write("✅ Requesting AI to Transform Data:")
    
    response = openai.ChatCompletion.create(
        model="gpt-4-turbo",
        messages=[{"role": "system", "content": "You are an AI that transforms raw CSV data into structured Excel format."},
                  {"role": "user", "content": prompt}],
        api_key=OPENAI_API_KEY
    )

    try:
        content = response["choices"][0]["message"]["content"]
        
        # Extract JSON from AI response
        json_start = content.find("[")
        json_end = content.rfind("]") + 1
        json_str = content[json_start:json_end]  # Extract JSON part

        # Parse JSON to dictionary list
        transformed_data = json.loads(json_str)
        st.write("✅ AI Transformed Data Preview:", transformed_data[:5])  # Show first 5 rows
        return transformed_data
    
    except Exception as e:
        st.error(f"⚠️ Error parsing AI-transformed data: {e}")
        return []

def populate_template(master_df, template_file):
    """Creates a new Excel file with AI-transformed data matching the template structure."""
    
    try:
        # Load the template
        template_wb = openpyxl.load_workbook(template_file)
        template_sheet = template_wb.active  
        template_headers = [cell.value for cell in template_sheet[1] if cell.value]  # Extract column names

        st.write("✅ Extracted Template Headers:", template_headers)

        # Get AI-transformed data
        transformed_data = get_ai_transformed_data(master_df, template_headers)
        st.write("✅ Extracted AI-Transformed Data:", transformed_data[:50])
        if not transformed_data:
            st.error("⚠️ AI failed to generate transformed data.")
            return None

        # Create a new workbook
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Sheet1"

        # Write headers
        new_sheet.append(template_headers)

        # Write transformed data
        for row in transformed_data:
            new_row = [row.get(col, "") for col in template_headers]  # Ensure all template columns exist
            new_sheet.append(new_row)

        # Save to BytesIO
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)

        st.write("✅ New Excel File Created with AI-transformed data!")
        return output

    except Exception as e:
        st.error(f"⚠️ Error creating new Excel file: {e}")
        return None
