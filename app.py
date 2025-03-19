import os
import pandas as pd
import streamlit as st
import openai
import json
from io import BytesIO
import openpyxl
from dotenv import load_dotenv
import requests

# Load API Key
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    st.error("⚠️ OpenAI API Key is missing! Set it in environment variables.")
    st.stop()

openai.api_key = OPENAI_API_KEY  # Set API key

### --------------- AGENT 1: Extract Context from CSV Files --------------- ###
def extract_csv_context(uploaded_files):
    """Reads multiple uploaded CSV files into Pandas DataFrames with debugging."""
    csv_data = {}

    for file in uploaded_files:
        try:
            df = pd.read_csv(file, encoding="utf-8", on_bad_lines="skip")
            csv_data[file.name] = df
            st.write(f"✅ Loaded {file.name}: {df.shape}")
        except Exception as e:
            st.error(f"⚠️ Error loading {file.name}: {e}")

    if not csv_data:
        st.error("⚠️ No valid CSV files were loaded.")
        return None

    return csv_data


### --------------- AGENT 2: Determine Merge Strategy --------------- ###
# def determine_merge_strategy(csv_data):
#     """Uses AI to decide the best merging strategy based on CSV context."""
#     prompt = f"Given the column names from multiple CSVs: {list(csv_data.keys())}, suggest a merging strategy."
#     response = openai.ChatCompletion.create(
#         model="gpt-4",
#         messages=[{"role": "user", "content": prompt}]
#     )
    
#     strategy = response["choices"][0]["message"]["content"]
#     st.write("🔍 Debug: AI Merge Strategy ->", strategy)
#     return strategy
def determine_merge_strategy(csv_data):
    """Uses AI to decide the best merging strategy based on CSV context and ensures proper formatting."""
    
    # Extract column names from CSV data
    csv_columns = {name: list(df.columns) for name, df in csv_data.items()}
    
    prompt = (
        "Given the following CSV column structures:\n"
        f"{json.dumps(csv_columns, indent=2)}\n\n"
        "Suggest a merging strategy. Respond in JSON format as either:\n"
        "1. {\"strategy\": \"Merge on common columns\"}\n"
        "2. {\"strategy\": \"Merge on key column\", \"key_column\": \"column_name\"}\n"
        "Ensure the key_column is present in all files if suggested."
    )

    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}]
        )

        # Extract and parse the response
        strategy_text = response["choices"][0]["message"]["content"]
        merge_strategy = json.loads(strategy_text)

        # Validate strategy format
        if not isinstance(merge_strategy, dict) or "strategy" not in merge_strategy:
            raise ValueError("Invalid strategy format received.")

        st.write("🔍 Debug: AI Merge Strategy ->", merge_strategy)
        return merge_strategy

    except json.JSONDecodeError:
        st.error("⚠️ Failed to parse AI response. Using fallback strategy.")
        return {"strategy": "Merge on common columns"}

    except Exception as e:
        st.error(f"⚠️ Error determining merge strategy: {e}")
        return {"strategy": "Merge on common columns"}  # Default fallback



# --------------- AGENT 3: Merge CSV Data --------------- ###

def merge_using_ai(csv_data, key_column):
    """Uses OpenAI API to intelligently merge CSVs on the given key column and group by it."""

    # Generate the AI prompt to guide the merge logic
    prompt = f"Merge the following CSV datasets using '{key_column}' as the key column. Handle missing values intelligently and resolve conflicts where possible. Ensure no duplicate entries. Group the data based on the {key_column}."

    # Send the prompt to OpenAI's API
    response = openai.ChatCompletion.create(
        model="gpt-4o",
        messages=[{"role": "system", "content": "You are an expert data analyst."},
                  {"role": "user", "content": prompt}]
    )

    # Get the merge instructions from the AI response
    merge_instructions = response["choices"][0]["message"]["content"].strip()
    st.write("🤖 AI Merge Instructions:", merge_instructions)

    # Merge the CSV datasets using the key column
    master_df = None
    for csv_name, df in csv_data.items():
        if master_df is None:
            master_df = df
        else:
            master_df = pd.merge(master_df, df, on=key_column, how='outer')  # Outer join to retain all rows

    # Handle missing values intelligently
    master_df = master_df.fillna('Missing')  # Example, can be customized based on AI's suggestion

    # Group the data based on the key_column
    grouped_df = master_df.groupby(key_column).agg('first').reset_index()  # Example aggregation: take the first value of each group

    # Display the grouped and merged dataframe
    st.write(grouped_df)

    return grouped_df

def merge_csv_data(csv_data, key_column, merge_strategy):
    """Merges multiple CSVs using OpenAI-determined method."""
    try:
        if merge_strategy == "Merge on key column":
            master_df = merge_using_ai(csv_data, key_column)
        else:
            master_df = pd.concat(csv_data.values(), ignore_index=True)
        
        st.write("✅ CSVs Merged Successfully using key column:", key_column)
        # st.write("✅ CSVs Merged Successfully: ", master_df.shape)
        return master_df
    
    except Exception as e:
        st.error(f"⚠️ Error merging CSV files: {e}")
        return None


### --------------- AGENT 4: Generate Data Mapping --------------- ###
def generate_data_map(master_df):
    """Creates a column mapping for AI to refine."""
    if master_df is None:
        st.error("⚠️ No master dataset found!")
        return None
    
    mapping = {col: "" for col in master_df.columns}
    st.write("🔍 Debug: Initial Mapping ->", mapping)
    return mapping


### --------------- AGENT 5: Extract Excel Template Context --------------- ###
def extract_template_context(template_file):
    """Reads the Excel template and extracts required fields."""
    try:
        df = pd.read_excel(template_file, engine="openpyxl")
        # df.columns = df.iloc[0]  # Set first row as header
        # df = df[1:].reset_index(drop=True)

        template_columns = list(df.columns)
        st.write("✅ Extracted Template Columns:", template_columns)
        return template_columns
    except Exception as e:
        st.error(f"⚠️ Error reading template: {e}")
        return None


### --------------- AGENT 6: AI-Powered Mapping --------------- ###


def ai_map_columns(master_df, template_columns):
    """Uses OpenAI to intelligently map CSV data to the Excel template in JSON format."""
    prompt = (
    f"Given these CSV columns: {list(master_df.columns)}, map them to these template columns: {template_columns}. "
    "Return the mapping as a **valid JSON object** with the format: "
    '{ "mappings": { "CSV_column_name": "Template_column_name", ... } }. '
    "Ensure that the CSV column names (from the given list) are used as **keys**, "
    "and the corresponding template column names (from the given list) are used as **values**. "
    "Do not swap the order. Only return a valid JSON object, no extra text."
)

    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    
    try:
        mapping = json.loads(response["choices"][0]["message"]["content"].strip())
        st.write("✅ AI Column Mapping:", mapping)
        # mapping_json = json.dumps(mapping, indent=4)  # Convert dict to JSON string with indentation
        # return mapping_json
        return mapping["mappings"]
    except json.JSONDecodeError:
        st.error("⚠️ Error: AI response is not valid JSON.")
        return None



### --------------- AGENT 7: Populate Excel Template --------------- ###

def populate_template(master_df, mapping, template_file):
    """Creates a new Excel file with mapped data from master_df while retaining template columns."""

    print("DEBUG: Checking mapping type...", type(mapping))
    if not isinstance(mapping, dict):
        raise TypeError(f"Expected dictionary for mapping, but got {type(mapping)}")

    try:
        # Load the original template (preserve the columns only)
        template_wb = openpyxl.load_workbook(template_file)
        template_sheet = template_wb.active  # Assuming single-sheet templates

        # Extract headers from the template (only column names, not data)
        template_headers = [cell.value for cell in template_sheet[1]]  # First row is header

        # Create a new workbook
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Sheet1"  # Name of the sheet (optional)

        # Write headers to the first row in the new sheet (using template headers)
        new_sheet.append(template_headers)  # Write headers to the first row

        # Populate the new sheet with data from master_df based on the mapping
        for _, row in master_df.iterrows():
            new_row = []
            for template_col in template_headers:
                # Find the corresponding column in master_df using mapping
                csv_col = next((key for key, value in mapping.items() if value == template_col), None)
                if csv_col and csv_col in master_df.columns:
                    new_row.append(row[csv_col])  # Add the value from master_df
                else:
                    new_row.append(None)  # If no match, add None (empty cell)
            new_sheet.append(new_row)  # Write the row to the sheet

        # Save the new workbook to a BytesIO object for streaming or saving
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)

        # Optional: Display the first few rows of the new data in Streamlit
        st.write(master_df.head(10))
        st.write("✅ New Excel File Created Successfully!")

        return output

    except Exception as e:
        print(f"⚠️ Error creating new Excel file: {e}")
        return None

# def get_api_data_for_column(col_value):
#     """Function to make an API call using the column value, returns processed data."""
#     # Example API call - replace with actual API and payload
#     api_url = "https://example.com/api"
#     params = {"value": col_value}
#     response = requests.get(api_url, params=params)
    
#     # Check if the API call is successful
#     if response.status_code == 200:
#         return response.json()  # Assuming the API returns JSON data
#     else:
#         return None

# def populate_template(master_df, mapping, template_file):
#     """Creates a new Excel file with mapped data from master_df while retaining template columns."""
    
#     print("DEBUG: Checking mapping type...", type(mapping))
#     if not isinstance(mapping, dict):
#         raise TypeError(f"Expected dictionary for mapping, but got {type(mapping)}")

#     try:
#         # Load the original template (preserve the columns only)
#         template_wb = openpyxl.load_workbook(template_file)
#         template_sheet = template_wb.active  # Assuming single-sheet templates

#         # Extract headers from the template (only column names, not data)
#         template_headers = [cell.value for cell in template_sheet[1]]  # First row is header

#         # Create a new workbook
#         new_wb = openpyxl.Workbook()
#         new_sheet = new_wb.active
#         new_sheet.title = "Sheet1"  # Name of the sheet (optional)

#         # Write headers to the first row in the new sheet (using template headers)
#         new_sheet.append(template_headers)  # Write headers to the first row

#         # Populate the new sheet with data from master_df based on the mapping
#         for _, row in master_df.iterrows():
#             new_row = []
#             for template_col in template_headers:
#                 # Find the corresponding column in master_df using mapping
#                 csv_col = next((key for key, value in mapping.items() if value == template_col), None)
                
#                 if csv_col and csv_col in master_df.columns:
#                     # Fetch the value from master_df
#                     col_value = row[csv_col]
                    
#                     # Make an API call to fetch data based on the column value (optional)
#                     api_data = get_api_data_for_column(col_value)
                    
#                     # You can modify how you handle the API data here
#                     if api_data:
#                         new_row.append(api_data)  # Or process the API data as needed
#                     else:
#                         new_row.append(col_value)  # Fallback to master_df value if no API response
#                 else:
#                     new_row.append(None)  # If no match, add None (empty cell)
#             new_sheet.append(new_row)  # Write the row to the sheet

#         # Save the new workbook to a BytesIO object for streaming or saving
#         output = BytesIO()
#         new_wb.save(output)
#         output.seek(0)

#         # Optional: Display the first few rows of the new data in Streamlit
#         st.write(master_df.head(10))
#         st.write("✅ New Excel File Created Successfully!")

#         return output

#     except Exception as e:
#         print(f"⚠️ Error creating new Excel file: {e}")
#         return None


### --------------- STREAMLIT UI --------------- ###
st.title("🔗 CSV-to-Excel Mapping App")

# Upload CSV files
uploaded_csvs = st.file_uploader("Upload CSV Files", type=["csv"], accept_multiple_files=True)

# Upload Excel template
uploaded_template = st.file_uploader("Upload Excel Template", type=["xlsx"])

# Process Button
if st.button("Process Files"):
    if not uploaded_csvs or not uploaded_template:
        st.error("⚠️ Please upload both CSV files and an Excel template.")
    else:
        # Agent 1: Extract CSV Data
        csv_dataframes = extract_csv_context(uploaded_csvs)
        
        if csv_dataframes:
            # Agent 2: Determine Merge Strategy
            merge_strategy = determine_merge_strategy(csv_dataframes)
            print("DEBUG: Merge Strategy →", merge_strategy)

            # Agent 3: Merge CSVs
            master_data = merge_csv_data(csv_dataframes, merge_strategy['key_column'], merge_strategy['strategy'])
            st.write("DEBUG: Master DataFrame Columns:", master_data.columns)
            st.write("DEBUG: Master DataFrame Sample Data:\n", master_data.head())
            if master_data is not None:
                # Agent 4: Generate Data Mapping
                initial_mapping = generate_data_map(master_data)

                # Agent 5: Extract Template Context
                template_fields = extract_template_context(uploaded_template)
                
                if template_fields:
                    # Agent 6: AI-Powered Mapping
                    refined_mapping = ai_map_columns(master_data, template_fields)
		    
                    # Assume `refined_mapping` is the string causing the issue

                print("Refined Mapping Before:",refined_mapping)
                if isinstance(refined_mapping, str):
                    try:
                        # refined_mapping = json.loads(refined_mapping["choices"][0]["message"]["content"].strip())

                        # print("Refined Mapping After:",refined_mapping)
                        refined_mapping = json.loads(refined_mapping)  # Convert to dictionary
                        st.write("✅ Successfully converted refined_mapping to dictionary.")
                    except json.JSONDecodeError as e:
                        print("⚠️ JSON Decode Error:", e)
                        raise ValueError("❌ Invalid JSON format. Please check refined_mapping.")
                
                if not isinstance(refined_mapping, dict):
                    st.error("⚠️ AI response is not a valid mapping dictionary.")
                    st.stop()


                # Agent 7: Populate Template
                final_file = populate_template(master_data, refined_mapping, uploaded_template)

                # Download Button
                if final_file:
                    st.download_button(
                        label="📥 Download Processed Excel",
                        data=final_file,
                        file_name="populated_template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
